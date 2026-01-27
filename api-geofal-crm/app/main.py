from __future__ import annotations
 
import io
import json # Added by user
import os
import re
import zipfile
from copy import copy
from datetime import date, datetime # Added datetime by user
from pathlib import Path # Corrected from 'from pathlib import os'
from typing import Any
 
import asyncio
import requests # Moved by user
from fastapi import FastAPI, HTTPException, Header, Response, Depends # Added Response, Depends; kept Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse # Kept StreamingResponse
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.worksheet.pagebreak import Break
from openpyxl.utils.cell import get_column_letter, range_boundaries
from pydantic import BaseModel, Field
import psycopg2
from psycopg2.extras import RealDictCursor

# Importar el nuevo exportador XML
from app.xlsx_direct_v2 import export_xlsx_direct
 
app = FastAPI(title="quotes-service")

load_dotenv(Path(__file__).resolve().parents[1] / ".env", override=False)

if (os.getenv("QUOTES_DATABASE_URL") or "").strip() == "":
    os.environ.pop("QUOTES_DATABASE_URL", None)


def _db_disabled() -> bool:
    return (os.getenv("QUOTES_DISABLE_DB") or "").strip().lower() in {"1", "true", "yes", "on"}


def _get_cors_origins() -> list[str]:
    raw = os.getenv("QUOTES_CORS_ORIGINS")
    if raw:
        raw = raw.strip()
        if raw == "*":
            return ["*"]
        return [o.strip() for o in raw.split(",") if o.strip()]
    return ["*"]
 
 
# Determine CORS origins
_origins = _get_cors_origins()
# If origins are set specifically, allow credentials. If it's "*", we cannot.
_allow_creds = "*" not in _origins and len(_origins) > 0

print(f"DEBUG: CORS Origins: {_origins}")
print(f"DEBUG: Allow Credentials: {_allow_creds}")

app.add_middleware(
    CORSMiddleware,
    allow_origins=_origins,
    allow_credentials=_allow_creds,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS", "PATCH"],
    allow_headers=["*"],
    expose_headers=["Content-Disposition"],
    max_age=3600,
)
 
 
@app.get("/")
@app.get("/health")
async def health_check():
    return {"status": "ok", "service": "quotes-service", "db": _has_database_url()}


@app.get("/debug-db")
async def debug_db():
    """Verify database connection and schema for troubleshooting"""
    if not _has_database_url():
        return {"error": "DATABASE_URL not set"}
    try:
        conn = _get_connection()
        with conn.cursor() as cur:
            cur.execute("SELECT version();")
            version = cur.fetchone()[0]
            
            cur.execute("SELECT table_name FROM information_schema.tables WHERE table_schema='public'")
            tables = [r[0] for r in cur.fetchall()]
            
            return {
                "status": "connected",
                "version": version,
                "tables": tables,
                "dsn_start": _get_database_url().split('@')[-1] if '@' in _get_database_url() else "local"
            }
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"DEBUG-DB Error: {e}\n{tb}")
        return {"status": "error", "message": str(e), "traceback": tb}
    finally:
        if 'conn' in locals() and conn:
            conn.close()


class QuoteItem(BaseModel):
    codigo: str
    descripcion: str
    norma: str | None = None
    acreditado: str | None = None
    costo_unitario: float = Field(ge=0)
    cantidad: float = Field(ge=0)
 
 
class QuoteExportRequest(BaseModel):
    cotizacion_numero: str | None = None
    fecha_emision: date | None = None
    fecha_solicitud: date | None = None
    cliente: str | None = None
    ruc: str | None = None
    contacto: str | None = None
    telefono_contacto: str | None = None
    correo: str | None = None
    proyecto: str | None = None
    ubicacion: str | None = None
    personal_comercial: str | None = None
    telefono_comercial: str | None = None
    correo_vendedor: str | None = None
    plazo_dias: int | None = None
    condicion_pago: str | None = None
    condiciones_ids: list[str] | None = Field(default_factory=list)
    include_igv: bool = True
    igv_rate: float = 0.18
    items: list[QuoteItem] = Field(default_factory=list)
    template_id: str | None = None
    user_id: str | None = None
    proyecto_id: str | None = None
 
 
class NextNumberResponse(BaseModel):
    year: int
    sequential: int
    token: str
 
 
# Mapping of template IDs to filenames
TEMPLATE_VARIANTS = {
    'V1': 'V1 - MUESTRA DE SUELO Y AGREGADO.xlsx',
    'V2': 'V2 - PROBETAS.xlsx',
    'V3': 'V3 - DENSIDAD DE CAMPO Y MUESTREO.xlsx',
    'V4': 'V4 - EXTRACCIÓN DE DIAMANTINA.xlsx',
    'V5': 'V5 - DIAMANTINA PARA PASES.xlsx',
    'V6': 'V6 - ALBAÑILERÍA.xlsx',
    'V7': 'V7 - VIGA BECKELMAN.xlsx',
    'V8': 'V8 - CONTROL DE CALIDAD DE CONCRETO FRESCO EN OBRA.xlsx',
}

def _get_template_path(template_id: str | None = None) -> Path:
    """Get template path based on template_id or default"""
    filename = TEMPLATE_VARIANTS.get(template_id, 'Formato-cotizacion.xlsx') if template_id else 'Formato-cotizacion.xlsx'
    
    # Try multiple possible locations
    possible_paths = [
        Path(__file__).resolve().parents[2] / filename,  # Local dev
        Path(__file__).resolve().parents[1] / filename,  # Docker context
        Path("/app") / filename,  # Docker absolute
        Path("/app/templates") / filename,  # Docker templates folder
    ]
    for p in possible_paths:
        if p.exists():
            print(f"Template found: {p}")
            return p
    print(f"Template not found, tried: {[str(p) for p in possible_paths]}")
    return possible_paths[0]  # Fallback


def _default_template_path() -> Path:
    return _get_template_path(None)
 
 
def _get_database_url() -> str:
    url = os.getenv("QUOTES_DATABASE_URL")
    if not url:
        raise RuntimeError("Missing QUOTES_DATABASE_URL env var")
    return url
 
 
def _has_database_url() -> bool:
    if _db_disabled():
        return False
    url = (os.getenv("QUOTES_DATABASE_URL") or "").strip()
    if url.startswith("http"):
        print("WARNING: QUOTES_DATABASE_URL starts with http. It should be a postgresql:// DSN.")
    return bool(url)


# Startup check to help users catch configuration errors in logs
_db_url = (os.getenv("QUOTES_DATABASE_URL") or "").strip()
_sb_url = (os.getenv("SUPABASE_URL") or "").strip()

if _db_url:
    print(f"INFO: Database connection configured (dsn starts with {_db_url[:10]}...)")
    if _db_url.startswith("http"):
        print("CRITICAL: QUOTES_DATABASE_URL is set to an HTTP URL! It must be a postgresql:// connection string.")

if _sb_url:
    print(f"INFO: Supabase URL configured: {_sb_url}")
    if _sb_url.startswith("postgresql://"):
        print("CRITICAL: SUPABASE_URL is set to a postgresql:// connection string! It must be an HTTP URL (e.g. https://db.geofal.com.pe).")


def _get_connection():
    """Get a new database connection."""
    dsn = _get_database_url()
    return psycopg2.connect(dsn)


def _upload_to_supabase_storage(file_data: io.BytesIO, bucket: str, path: str) -> str | None:
    """Uploads a file to Supabase Storage and returns the public URL or partial path."""
    
    url = os.getenv("SUPABASE_URL")
    key = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
    
    if not url or not key:
        print("Supabase Storage credentials not found")
        return None
        
    storage_url = f"{url.rstrip('/')}/storage/v1/object/{bucket}/{path}"
    
    file_data.seek(0)
    try:
        resp = requests.post(
            storage_url,
            headers={
                "Authorization": f"Bearer {key}",
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "x-upsert": "true"
            },
            data=file_data.read()
        )
        if resp.status_code == 200:
            print(f"File uploaded to {bucket}/{path}")
            return f"{bucket}/{path}"
        else:
            print(f"Storage upload failed: {resp.status_code} - {resp.text}")
            return None
    except Exception as e:
        print(f"Error uploading to storage: {e}")
        return None
 
 
def _ensure_sequence_table() -> None:
    try:
        dsn = _get_database_url()
        with psycopg2.connect(dsn) as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS quote_sequences (
                      year INTEGER PRIMARY KEY,
                      last_value INTEGER NOT NULL
                    );
                    """
                )
    except Exception as e:
        import traceback
        print(f"Error in _ensure_sequence_table: {e}")
        traceback.print_exc()
        raise
 
 
def _next_quote_sequential(year: int) -> int:
    dsn = _get_database_url()
    try:
        with psycopg2.connect(dsn) as conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute("SELECT year, last_value FROM quote_sequences WHERE year = %s FOR UPDATE", (year,))
                row = cur.fetchone()
                if row is None:
                    cur.execute("INSERT INTO quote_sequences (year, last_value) VALUES (%s, %s)", (year, 0))
                    last_value = 0
                else:
                    last_value = int(row["last_value"])

                next_value = last_value + 1
                cur.execute("UPDATE quote_sequences SET last_value = %s WHERE year = %s", (next_value, year))
                return next_value
    except Exception as exc:
        raise RuntimeError("Failed to connect to QUOTES_DATABASE_URL") from exc
 
 
def _format_quote_token(sequential: int, year: int) -> str:
    return f"{sequential:03d}-{str(year)[-2:]}"
 
 
def _copy_row_format(ws: Any, src_row: int, dst_row: int, *, min_col: int = 1, max_col: int = 60) -> None:
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for col in range(min_col, max_col + 1):
        src_cell = ws.cell(row=src_row, column=col)
        dst_cell = ws.cell(row=dst_row, column=col)
        if src_cell.has_style:
            dst_cell._style = copy(src_cell._style)
        dst_cell.number_format = src_cell.number_format
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.comment = None


def _shift_range_rows(range_ref: str, *, insert_at_row: int, delta: int) -> str:
    min_col, min_row, max_col, max_row = range_boundaries(range_ref)
    if min_row >= insert_at_row:
        min_row += delta
        max_row += delta
    return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


def _restore_merged_cells(ws: Any, merged_ranges: list[str], *, insert_at_row: int, delta: int) -> None:
    for r in list(ws.merged_cells.ranges):
        try:
            ws.unmerge_cells(str(r))
        except Exception:
            continue

    for r in merged_ranges:
        new_ref = _shift_range_rows(r, insert_at_row=insert_at_row, delta=delta)
        try:
            ws.merge_cells(new_ref)
        except ValueError:
            continue


def _restore_print_area(ws: Any, print_area: Any, *, insert_at_row: int, delta: int) -> None:
    if not print_area:
        return

    raw = print_area
    if isinstance(raw, list):
        raw = raw[0] if raw else None
    if not raw or not isinstance(raw, str):
        return

    try:
        min_col, min_row, max_col, max_row = range_boundaries(raw)
    except Exception:
        return

    if max_row >= insert_at_row:
        max_row += delta
    ws.print_area = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"


def _force_merge_b_to_n(ws: Any, row: int) -> None:
    target = f"B{row}:N{row}"

    for r in list(ws.merged_cells.ranges):
        try:
            min_col, min_row, max_col, max_row = range_boundaries(str(r))
        except Exception:
            continue

        if min_row <= row <= max_row and not (max_col < 2 or min_col > 14):
            try:
                ws.unmerge_cells(str(r))
            except Exception:
                continue

    try:
        ws.merge_cells(target)
    except ValueError:
        return


def _force_merge_range(ws: Any, *, row: int, min_col: int, max_col: int) -> None:
    target = f"{get_column_letter(min_col)}{row}:{get_column_letter(max_col)}{row}"

    for r in list(ws.merged_cells.ranges):
        try:
            r_min_col, r_min_row, r_max_col, r_max_row = range_boundaries(str(r))
        except Exception:
            continue

        if r_min_row <= row <= r_max_row and not (r_max_col < min_col or r_min_col > max_col):
            try:
                ws.unmerge_cells(str(r))
            except Exception:
                continue

    try:
        ws.merge_cells(target)
    except ValueError:
        return


def _find_row_by_text(ws: Any, text: str, *, max_rows: int = 200, max_cols: int = 20) -> int | None:
    needle = text.strip().lower()
    for r in range(1, min(max_rows, ws.max_row) + 1):
        for c in range(1, min(max_cols, ws.max_column) + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            if isinstance(v, str) and needle in v.strip().lower():
                return r
    return None


def _snapshot_row_style(ws: Any, *, row: int, min_col: int, max_col: int) -> dict[int, dict[str, Any]]:
    snap: dict[int, dict[str, Any]] = {}
    snap[0] = {"height": ws.row_dimensions[row].height}
    for c in range(min_col, max_col + 1):
        cell = ws.cell(row=row, column=c)
        snap[c] = {
            "_style": copy(cell._style),
            "number_format": cell.number_format,
            "alignment": copy(cell.alignment),
            "font": copy(cell.font),
            "border": copy(cell.border),
            "fill": copy(cell.fill),
            "protection": copy(cell.protection),
        }
    return snap


def _snapshot_row_merges(ws: Any, *, row: int) -> list[str]:
    merges: list[str] = []
    for r in ws.merged_cells.ranges:
        try:
            min_col, min_row, max_col, max_row = range_boundaries(str(r))
        except Exception:
            continue
        if min_row <= row <= max_row:
            merges.append(str(r))
    return merges


def _apply_row_merges(ws: Any, *, row: int, merges: list[str], insert_at_row: int, delta: int) -> None:
    if not merges:
        return

    for existing in list(ws.merged_cells.ranges):
        try:
            _, min_row, _, max_row = range_boundaries(str(existing))
        except Exception:
            continue
        if min_row <= row <= max_row:
            try:
                ws.unmerge_cells(str(existing))
            except Exception:
                continue

    for rng in merges:
        try:
            min_col, min_row, max_col, max_row = range_boundaries(rng)
        except Exception:
            continue

        if delta > 0:
            if min_row >= insert_at_row:
                min_row += delta
                max_row += delta
            elif max_row >= insert_at_row:
                max_row += delta

        target = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
        try:
            ws.merge_cells(target)
        except ValueError:
            continue


def _apply_row_style(ws: Any, *, row: int, min_col: int, max_col: int, snap: dict[int, dict[str, Any]]) -> None:
    if 0 in snap and "height" in snap[0]:
        ws.row_dimensions[row].height = snap[0]["height"]

    for c in range(min_col, max_col + 1):
        if c not in snap:
            continue
        dst = ws.cell(row=row, column=c)
        s = snap[c]
        dst._style = copy(s["_style"])
        dst.number_format = s["number_format"]
        dst.alignment = copy(s["alignment"])
        dst.font = copy(s["font"])
        dst.border = copy(s["border"])
        dst.fill = copy(s["fill"])
        dst.protection = copy(s["protection"])


def _set_cell(ws: Any, addr: str, value: Any) -> None:
    if value is None:
        return

    cell = ws[addr]
    try:
        cell.value = value
        return
    except AttributeError:
        pass

    try:
        col, row = ws[addr].column, ws[addr].row
    except Exception:
        ws[addr].value = value
        return

    for r in ws.merged_cells.ranges:
        try:
            min_col, min_row, max_col, max_row = range_boundaries(str(r))
        except Exception:
            continue
        if min_col <= col <= max_col and min_row <= row <= max_row:
            top_left = ws.cell(row=min_row, column=min_col)
            top_left.value = value
            return

    ws[addr].value = value


def _apply_quote_number(ws: Any, addr: str, cotizacion_numero: str | None, fecha_emision: date | None) -> None:
    if not cotizacion_numero and not fecha_emision:
        return

    current = ws[addr].value
    if not isinstance(current, str):
        current = ""

    if fecha_emision is None:
        fecha_emision = date.today()

    year_suffix = str(fecha_emision.year)[-2:]
    numero = cotizacion_numero or "000"
    token = f"{numero}-{year_suffix}"

    if "XXX-XX" in current:
        ws[addr].value = current.replace("XXX-XX", token)
        return

    if re.search(r"XXX-\d{2}", current):
        ws[addr].value = re.sub(r"XXX-\d{2}", token, current)
        return

    ws[addr].value = re.sub(r"\b\d{1,6}-\d{2}\b", token, current) or token


def _shift_drawing_xml(data: bytes, *, start_row0: int, delta: int) -> bytes:
    """Desplaza anchors en xl/drawings/*.xml.

    En DrawingML, la posición de los shapes/imágenes viene en <xdr:from>/<xdr:to>
    con <xdr:row>. Solo debemos desplazar esos valores (0-index) cuando se insertan filas.
    """

    if delta <= 0:
        return data

    try:
        text = data.decode("utf-8")
    except UnicodeDecodeError:
        return data

    def repl(m: re.Match[str]) -> str:
        v = int(m.group(2))
        if v >= start_row0:
            v += delta
        return f"{m.group(1)}{v}{m.group(3)}"

    # Solo desplaza anchors de DrawingML (xdr:row). No tocar otros <row>.
    # Para evitar desplazar drawings de otras hojas del template, solo aplicamos
    # shift si el XML contiene anchors en rangos razonables (<=200), típicos de la
    # hoja de cotización.
    row_vals = [int(v) for v in re.findall(r"<xdr:row>(\d+)</xdr:row>", text)]
    has_xdr_rows = bool(row_vals)
    if not row_vals:
        row_vals = [int(v) for v in re.findall(r"<[A-Za-z0-9_]+:row>(\d+)</[A-Za-z0-9_]+:row>", text)]

    if not any(start_row0 <= v <= 200 for v in row_vals):
        return data

    # Evitar doble desplazamiento: si ya tenemos <xdr:row>, solo desplazamos esos.
    if has_xdr_rows:
        text = re.sub(r"(<xdr:row>)(\d+)(</xdr:row>)", repl, text)
    else:
        text = re.sub(r"(<[A-Za-z0-9_]+:row>)(\d+)(</[A-Za-z0-9_]+:row>)", repl, text)

    return text.encode("utf-8")


def _shift_vml(data: bytes, *, start_row0: int, delta: int) -> bytes:
    if delta <= 0:
        return data
    try:
        text = data.decode("utf-8")
    except UnicodeDecodeError:
        try:
            text = data.decode("latin-1")
        except Exception:
            return data

    def repl_row(m: re.Match[str]) -> str:
        v = int(m.group(1))
        if v >= start_row0:
            v += delta
        return f"{m.group(0).split('>')[0]}>" + str(v) + f"</{m.group(0).split('</')[1]}"


    # <x:Row>n</x:Row>
    text = re.sub(r"(<x:Row>)(\d+)(</x:Row>)", lambda m: f"{m.group(1)}{int(m.group(2)) + (delta if int(m.group(2)) >= start_row0 else 0)}{m.group(3)}", text)

    # <x:Anchor>col,dx,row,dy,col2,dx2,row2,dy2</x:Anchor>
    def repl_anchor(m: re.Match[str]) -> str:
        parts = [p.strip() for p in m.group(2).split(",")]
        try:
            nums = [int(p) for p in parts]
        except Exception:
            return m.group(0)

        if len(nums) >= 8:
            if nums[2] >= start_row0:
                nums[2] += delta
            if nums[6] >= start_row0:
                nums[6] += delta
            new_val = ",".join(str(n) for n in nums)
            return f"{m.group(1)}{new_val}{m.group(3)}"
        return m.group(0)

    text = re.sub(r"(<x:Anchor>)([^<]+)(</x:Anchor>)", repl_anchor, text)

    # Evitar desplazar VML de otras hojas con anchors absurdamente grandes.
    anchors = re.findall(r"<x:Anchor>([^<]+)</x:Anchor>", text)
    ok = False
    for a in anchors:
        parts = [p.strip() for p in a.split(",")]
        if len(parts) < 8:
            continue
        try:
            r1 = int(parts[2])
            r2 = int(parts[6])
        except Exception:
            continue
        if (start_row0 <= r1 <= 200) or (start_row0 <= r2 <= 200):
            ok = True
            break

    if not ok:
        return data

    try:
        return text.encode("utf-8")
    except Exception:
        return data


def _preserve_template_assets(template_path: Path, generated: io.BytesIO, *, insert_at_row: int, delta: int) -> io.BytesIO:
    """openpyxl no preserva ciertos artefactos (logos/headers/footers) que vienen como
    drawings/VML. Para mantener el look del template, copiamos esas partes del ZIP
    original al XLSX generado.
    """

    generated.seek(0)
    out = io.BytesIO()

    with zipfile.ZipFile(template_path, "r") as ztpl:
        with zipfile.ZipFile(generated, "r") as zgen:
            with zipfile.ZipFile(out, "w", compression=zipfile.ZIP_DEFLATED) as zout:
                names_tpl = set(ztpl.namelist())

                def should_take_from_template(name: str) -> bool:
                    # Solo copiar media y drawings del template
                    # NO copiar relaciones ni Content_Types para evitar corrupción
                    if name.startswith("xl/media/"):
                        return True
                    if name.startswith("xl/drawings/"):
                        return True
                    # NO copiar xl/worksheets/_rels/ - dejar que openpyxl lo maneje
                    # NO copiar xl/_rels/ - dejar que openpyxl lo maneje
                    # NO copiar [Content_Types].xml - dejar que openpyxl lo maneje
                    return False

                tpl_override = {n for n in names_tpl if should_take_from_template(n)}
                gen_names = set(zgen.namelist())

                start_row0 = max(0, insert_at_row - 2)

                # Copiar archivos del generado primero
                for name in zgen.namelist():
                    if name in tpl_override and name in names_tpl:
                        data = ztpl.read(name)
                        if name.startswith("xl/drawings/") and name.endswith(".xml"):
                            data = _shift_drawing_xml(data, start_row0=start_row0, delta=delta)
                        if name.startswith("xl/drawings/") and name.endswith(".vml"):
                            data = _shift_vml(data, start_row0=start_row0, delta=delta)
                    else:
                        data = zgen.read(name)
                    zout.writestr(name, data)

                # Copiar archivos del template que no están en el generado
                for name in tpl_override:
                    if name not in gen_names and name in names_tpl:
                        data = ztpl.read(name)
                        if name.startswith("xl/drawings/") and name.endswith(".xml"):
                            data = _shift_drawing_xml(data, start_row0=start_row0, delta=delta)
                        if name.startswith("xl/drawings/") and name.endswith(".vml"):
                            data = _shift_vml(data, start_row0=start_row0, delta=delta)
                        zout.writestr(name, data)

    out.seek(0)
    return out


def _export_xlsx(payload: QuoteExportRequest) -> io.BytesIO:
    """
    Exporta XLSX modificando el XML del template directamente.
    Preserva todos los estilos, logos, footers y márgenes.
    """
    template_path = _get_template_path(payload.template_id)
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    # Generar número de cotización
    fecha_emision = payload.fecha_emision or date.today()
    cotizacion_numero = payload.cotizacion_numero
    if not cotizacion_numero:
        if _has_database_url():
            try:
                sequential = _next_quote_sequential(fecha_emision.year)
                cotizacion_numero = f"{sequential:03d}"
            except Exception:
                cotizacion_numero = "000"
        else:
            cotizacion_numero = "000"

    # Preparar datos para el exportador XML
    export_data = {
        'cotizacion_numero': cotizacion_numero,
        'fecha_emision': fecha_emision,
        'cliente': payload.cliente or '',
        'ruc': payload.ruc or '',
        'contacto': payload.contacto or '',
        'telefono': payload.telefono_contacto or '',
        'email': payload.correo or '',
        'correo': payload.correo_vendedor or payload.correo or '',
        'fecha_solicitud': payload.fecha_solicitud,
        'proyecto': payload.proyecto or '',
        'ubicacion': payload.ubicacion or '',
        'personal_comercial': payload.personal_comercial or '',
        'telefono_comercial': payload.telefono_comercial or '',
        'plazo_dias': payload.plazo_dias or 0,
        'condicion_pago': payload.condicion_pago or '',
        'condiciones_ids': payload.condiciones_ids or [],
        'items': [
            {
                'codigo': item.codigo,
                'descripcion': item.descripcion,
                'norma': item.norma,
                'acreditado': item.acreditado,
                'costo_unitario': item.costo_unitario,
                'cantidad': item.cantidad,
            }
            for item in payload.items
        ],
        'include_igv': payload.include_igv,
        'igv_rate': payload.igv_rate,
    }
    
    # Si hay condiciones seleccionadas, obtener sus textos de la BD
    if export_data['condiciones_ids']:
        try:
            conn = _get_connection()
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                ids_placeholder = ','.join(['%s'] * len(export_data['condiciones_ids']))
                cur.execute(f"""
                    SELECT texto FROM condiciones_especificas
                    WHERE id IN ({ids_placeholder}) AND activo = true
                    ORDER BY orden ASC
                """, export_data['condiciones_ids'])
                condiciones_textos = [row['texto'] for row in cur.fetchall()]
                export_data['condiciones_textos'] = condiciones_textos
                print(f"DEBUG MAIN: {len(condiciones_textos)} condiciones cargadas")
        except Exception as e:
            print(f"Warning: Could not load condiciones: {e}")
            export_data['condiciones_textos'] = []
        finally:
            if 'conn' in locals() and conn:
                conn.close()
    else:
        export_data['condiciones_textos'] = []
    
    print(f"DEBUG MAIN: plazo_dias={export_data['plazo_dias']}, condicion_pago={export_data['condicion_pago']}")
    print(f"DEBUG MAIN: correo={export_data['correo']}, telefono_comercial={export_data['telefono_comercial']}")

    # Usar el nuevo exportador que modifica XML directamente
    return export_xlsx_direct(str(template_path), export_data)


# Carpeta para guardar cotizaciones generadas
QUOTES_FOLDER = Path(__file__).resolve().parents[1] / "cotizaciones"
QUOTES_FOLDER.mkdir(exist_ok=True)


def _save_quote_to_folder(xlsx_bytes: io.BytesIO, cotizacion_numero: str, year: int, cliente: str) -> Path:
    """Guarda la cotización en la carpeta y retorna la ruta"""
    # Crear subcarpeta por año
    year_folder = QUOTES_FOLDER / str(year)
    year_folder.mkdir(exist_ok=True)
    
    # Nombre del archivo: COT-2026-001_Cliente.xlsx
    safe_cliente = re.sub(r'[^\w\s-]', '', cliente)[:30].strip().replace(' ', '_')
    filename = f"COT-{year}-{cotizacion_numero}_{safe_cliente}.xlsx"
    filepath = year_folder / filename
    
    # Guardar archivo
    xlsx_bytes.seek(0)
    with open(filepath, 'wb') as f:
        f.write(xlsx_bytes.read())
    
    return filepath


def _register_quote_in_db(cotizacion_numero: str, year: int, cliente: str, filepath: str, payload: QuoteExportRequest, object_key: str = None):
    """Registra la cotización en la base de datos"""
    if not _has_database_url():
        print("DB: No database URL configured")
        return None
    
    print(f"DB: Connecting to database... object_key={object_key}")
    conn = _get_connection()
    try:
        with conn.cursor() as cur:
            print(f"DB: Executing INSERT for {cotizacion_numero}-{year}")
            # Calcular total
            total = sum(item.costo_unitario * item.cantidad for item in payload.items)
            if payload.include_igv:
                total *= (1 + payload.igv_rate)
            
            # Insertar registro
            fecha_emision_val = payload.fecha_emision or date.today()
            fecha_solicitud_val = payload.fecha_solicitud or date.today()
            
            # Preparar items como JSON
            import json
            items_json = json.dumps([
                {
                    'codigo': item.codigo,
                    'descripcion': item.descripcion,
                    'norma': item.norma,
                    'acreditado': item.acreditado,
                    'costo_unitario': float(item.costo_unitario),
                    'cantidad': float(item.cantidad),
                }
                for item in payload.items
            ])
            
            subtotal = sum(item.costo_unitario * item.cantidad for item in payload.items)
            igv_amount = subtotal * payload.igv_rate if payload.include_igv else 0
            
            # Count items
            items_count = len(payload.items)
            template_id = payload.template_id or 'V1'
            
            # vendedor_nombre is the personal_comercial field
            vendedor_nombre = payload.personal_comercial or ''
            
            # user_id for user_created field
            user_id = payload.user_id if payload.user_id else None
            
            # Ensure proyecto_id is None if empty string or not provided
            proyecto_id = payload.proyecto_id if payload.proyecto_id and payload.proyecto_id.strip() else None

            # vendedor_id - if we have a user_id we use it as the seller ID
            # In the current DB schema it seems to be expected as a string/UUID
            vendedor_id = user_id
            
            cur.execute("""
                INSERT INTO cotizaciones (
                    numero, year, cliente_nombre, cliente_ruc, cliente_contacto, 
                    cliente_telefono, cliente_email, proyecto, ubicacion,
                    personal_comercial, telefono_comercial, fecha_solicitud, fecha_emision,
                    subtotal, igv, total, include_igv, estado, moneda, archivo_path, items_json,
                    template_id, items_count, vendedor_nombre, user_created, proyecto_id, vendedor_id, object_key
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (year, numero) DO UPDATE SET
                    cliente_nombre = EXCLUDED.cliente_nombre,
                    cliente_ruc = EXCLUDED.cliente_ruc,
                    cliente_contacto = EXCLUDED.cliente_contacto,
                    cliente_telefono = EXCLUDED.cliente_telefono,
                    cliente_email = EXCLUDED.cliente_email,
                    proyecto = EXCLUDED.proyecto,
                    ubicacion = EXCLUDED.ubicacion,
                    personal_comercial = EXCLUDED.personal_comercial,
                    telefono_comercial = EXCLUDED.telefono_comercial,
                    fecha_solicitud = EXCLUDED.fecha_solicitud,
                    fecha_emision = EXCLUDED.fecha_emision,
                    subtotal = EXCLUDED.subtotal,
                    igv = EXCLUDED.igv,
                    total = EXCLUDED.total,
                    include_igv = EXCLUDED.include_igv,
                    archivo_path = EXCLUDED.archivo_path,
                    items_json = EXCLUDED.items_json,
                    template_id = EXCLUDED.template_id,
                    items_count = EXCLUDED.items_count,
                    vendedor_nombre = EXCLUDED.vendedor_nombre,
                    proyecto_id = EXCLUDED.proyecto_id,
                    vendedor_id = EXCLUDED.vendedor_id,
                    object_key = EXCLUDED.object_key,
                    updated_at = CURRENT_TIMESTAMP
                RETURNING id
            """, (
                cotizacion_numero, year, payload.cliente, payload.ruc, payload.contacto,
                payload.telefono_contacto, payload.correo, payload.proyecto, payload.ubicacion,
                payload.personal_comercial, payload.telefono_comercial, fecha_solicitud_val, fecha_emision_val,
                subtotal, igv_amount, total, payload.include_igv, 'borrador', 'PEN', 
                str(filepath), items_json, template_id, items_count, vendedor_nombre, 
                user_id, proyecto_id, vendedor_id, object_key
            ))
            
            result = cur.fetchone()
            conn.commit()
            return result[0] if result else None
    except Exception as e:
        conn.rollback()
        raise e
    finally:
        conn.close()


@app.post("/export")
async def export_quote(payload: QuoteExportRequest) -> Response:
    try:
        xlsx_bytes = _export_xlsx(payload)
        
        # Obtener número de cotización para guardar
        year = (payload.fecha_emision or date.today()).year
        cotizacion_numero = payload.cotizacion_numero or "000"
        
        # Guardar en carpeta
        xlsx_bytes.seek(0)
        filepath = _save_quote_to_folder(
            io.BytesIO(xlsx_bytes.read()), 
            cotizacion_numero, 
            year, 
            payload.cliente or "SinCliente"
        )
        
        # Registrar en base de datos y subir a storage
        try:
            # Generar el cloud_path para Supabase Storage
            cloud_path = f"{year}/COT-{year}-{cotizacion_numero}-{payload.cliente or 'S-N'}.xlsx"
            
            # Registrar en DB con el object_key
            _register_quote_in_db(cotizacion_numero, year, payload.cliente, str(filepath), payload, object_key=cloud_path)
            print(f"Quote {cotizacion_numero}-{year} saved to database with object_key={cloud_path}")
            
            # Subir a Supabase Storage
            xlsx_bytes.seek(0)
            _upload_to_supabase_storage(xlsx_bytes, "cotizaciones", cloud_path)
            
        except Exception as db_error:
            import traceback
            print(f"DB save error: {db_error}")
            print(f"Traceback: {traceback.format_exc()}")
        
        # Retornar archivo
        xlsx_bytes.seek(0)
        return Response(
            content=xlsx_bytes.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="COT-{year}-{cotizacion_numero}.xlsx"'},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/export/xlsx")
async def export_quote_xlsx(payload: QuoteExportRequest) -> Response:
    """Alias para /export - usado por el frontend"""
    return await export_quote(payload)


@app.get("/quotes")
async def list_quotes(year: int = None, limit: int = 50):
    """Lista las cotizaciones guardadas"""
    quotes = []
    
    # Si hay base de datos, obtener de ahí
    if _has_database_url():
        conn = _get_connection()
        try:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                if year:
                    cur.execute("""
                        SELECT id, numero, year, cliente, ruc, proyecto, total, estado, moneda, fecha_emision, filepath, created_at
                        FROM cotizaciones
                        WHERE year = %s
                        ORDER BY created_at DESC
                        LIMIT %s
                    """, (year, limit))
                else:
                    cur.execute("""
                        SELECT id, numero, year, cliente, ruc, proyecto, total, estado, moneda, fecha_emision, filepath, created_at
                        FROM cotizaciones
                        ORDER BY created_at DESC
                        LIMIT %s
                    """, (limit,))
                quotes = [dict(row) for row in cur.fetchall()]
        finally:
            conn.close()
    else:
        # Listar desde carpeta
        target_year = year or date.today().year
        year_folder = QUOTES_FOLDER / str(target_year)
        if year_folder.exists():
            for f in sorted(year_folder.glob("*.xlsx"), reverse=True)[:limit]:
                quotes.append({
                    "filename": f.name,
                    "filepath": str(f),
                    "year": target_year,
                    "created_at": f.stat().st_mtime
                })
    
    return {"quotes": quotes, "total": len(quotes)}


@app.get("/quotes/{quote_id}/download")
async def download_quote(quote_id: int):
    """Descarga una cotización guardada por ID"""
    if not _has_database_url():
        raise HTTPException(status_code=400, detail="Database not configured")
    
    conn = _get_connection()
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("SELECT filepath FROM cotizaciones WHERE id = %s", (quote_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Quote not found")
            
            filepath = Path(row['filepath'])
            if not filepath.exists():
                raise HTTPException(status_code=404, detail="File not found")
            
            with open(filepath, 'rb') as f:
                content = f.read()
            
            return Response(
                content=content,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{filepath.name}"'},
            )
    finally:
        conn.close()


@app.delete("/quotes/{quote_id}")
async def delete_quote(quote_id: int):
    """Elimina una cotización por ID"""
    if not _has_database_url():
        raise HTTPException(status_code=400, detail="Database not configured")
    
    conn = _get_connection()
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            # Obtener filepath antes de eliminar
            cur.execute("SELECT filepath FROM cotizaciones WHERE id = %s", (quote_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Quote not found")
            
            # Eliminar archivo físico
            filepath = Path(row['filepath'])
            if filepath.exists():
                filepath.unlink()
            
            # Eliminar registro de base de datos
            cur.execute("DELETE FROM cotizaciones WHERE id = %s", (quote_id,))
            conn.commit()
            
            return {"success": True, "message": "Quote deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@app.get("/quotes/{quote_id}")
async def get_quote_by_id(quote_id: int):
    """Obtiene los detalles completos de una cotización para editar"""
    if not _has_database_url():
        raise HTTPException(status_code=400, detail="Database not configured")
    
    conn = _get_connection()
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT id, numero, year, cliente, ruc, contacto, telefono_contacto, correo,
                       proyecto, ubicacion, personal_comercial, telefono_comercial, correo_vendedor,
                       plazo_dias, condicion_pago, items_json, condiciones_ids,
                       total, estado, moneda, fecha_emision, fecha_solicitud, cliente_id, proyecto_id,
                       include_igv, created_at, updated_at
                FROM cotizaciones
                WHERE id = %s
            """, (quote_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Quote not found")
            
            return {"data": dict(row), "success": True}
    finally:
        conn.close()


@app.put("/quotes/{quote_id}")
async def update_quote(quote_id: int, payload: QuoteExportRequest):
    """Actualiza una cotización existente y regenera el archivo Excel"""
    if not _has_database_url():
        raise HTTPException(status_code=400, detail="Database not configured")
    
    conn = _get_connection()
    try:
        # Verificar que la cotización existe
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("SELECT numero, year, filepath, object_key FROM cotizaciones WHERE id = %s", (quote_id,))
            existing = cur.fetchone()
            if not existing:
                raise HTTPException(status_code=404, detail="Quote not found")
        
        # Obtener textos de condiciones si hay IDs
        condiciones_textos = []
        if payload.condiciones_ids and len(payload.condiciones_ids) > 0:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                placeholders = ','.join(['%s'] * len(payload.condiciones_ids))
                cur.execute(f"""
                    SELECT texto FROM condiciones_especificas
                    WHERE id = ANY(%s) AND activo = true
                    ORDER BY orden
                """, (payload.condiciones_ids,))
                condiciones_textos = [row['texto'] for row in cur.fetchall()]
        
        # Generar nuevo Excel
        year = existing['year']
        cotizacion_numero = existing['numero']
        
        payload_dict = {
            "cotizacion_numero": cotizacion_numero,
            "fecha_emision": payload.fecha_emision or datetime.now().strftime("%Y-%m-%d"),
            "fecha_solicitud": payload.fecha_solicitud or datetime.now().strftime("%Y-%m-%d"),
            "cliente": payload.cliente or "",
            "ruc": payload.ruc or "",
            "contacto": payload.contacto or "",
            "telefono_contacto": payload.telefono_contacto or "",
            "correo": payload.correo or "",
            "proyecto": payload.proyecto or "",
            "ubicacion": payload.ubicacion or "",
            "personal_comercial": payload.personal_comercial or "",
            "telefono_comercial": payload.telefono_comercial or "",
            "correo_vendedor": payload.correo_vendedor or "",
            "plazo_dias": payload.plazo_dias or 0,
            "condicion_pago": payload.condicion_pago or "",
            "condiciones_textos": condiciones_textos,
            "include_igv": payload.include_igv,
            "igv_rate": payload.igv_rate,
            "items": [
                {
                    "codigo": it.codigo or "",
                    "descripcion": it.descripcion or "",
                    "norma": it.norma or "",
                    "acreditado": it.acreditado or "NO",
                    "cantidad": it.cantidad,
                    "costo_unitario": it.costo_unitario,
                }
                for it in payload.items
            ],
        }
        
        # Exportar Excel actualizado
        xlsx_bytes = export_xlsx_direct(payload_dict)
        
        # Actualizar archivo físico
        if existing['filepath']:
            filepath = Path(existing['filepath'])
            if filepath.exists():
                filepath.unlink()
        
        # Guardar nuevo archivo
        year_folder = QUOTES_FOLDER / str(year)
        year_folder.mkdir(parents=True, exist_ok=True)
        filepath = year_folder / f"COT-{year}-{cotizacion_numero}.xlsx"
        
        xlsx_bytes.seek(0)
        with open(filepath, "wb") as f:
            f.write(xlsx_bytes.read())
        
        # Actualizar en BD
        items_json = json.dumps([{
            "codigo": it.codigo,
            "descripcion": it.descripcion,
            "norma": it.norma,
            "acreditado": it.acreditado,
            "cantidad": it.cantidad,
            "costo_unitario": it.costo_unitario,
            "total": it.cantidad * it.costo_unitario
        } for it in payload.items], ensure_ascii=False)
        
        total = sum(it.cantidad * it.costo_unitario for it in payload.items)
        if payload.include_igv:
            total *= (1 + payload.igv_rate)
        
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE cotizaciones
                SET cliente = %s, ruc = %s, contacto = %s, telefono_contacto = %s, correo = %s,
                    proyecto = %s, ubicacion = %s, personal_comercial = %s, telefono_comercial = %s,
                    correo_vendedor = %s, plazo_dias = %s, condicion_pago = %s,
                    items_json = %s, condiciones_ids = %s, total = %s,
                    fecha_emision = %s, fecha_solicitud = %s, filepath = %s,
                    include_igv = %s, updated_at = NOW(),
                    cliente_id = %s, proyecto_id = %s
                WHERE id = %s
            """, (
                payload.cliente, payload.ruc, payload.contacto, payload.telefono_contacto, payload.correo,
                payload.proyecto, payload.ubicacion, payload.personal_comercial, payload.telefono_comercial,
                payload.correo_vendedor, payload.plazo_dias, payload.condicion_pago,
                items_json, payload.condiciones_ids, total,
                payload.fecha_emision, payload.fecha_solicitud, str(filepath),
                payload.include_igv,
                payload.cliente_id, payload.proyecto_id,
                quote_id
            ))
            conn.commit()
        
        # Subir a Supabase Storage si existe object_key
        if existing.get('object_key'):
            try:
                xlsx_bytes.seek(0)
                _upload_to_supabase_storage(xlsx_bytes, "cotizaciones", existing['object_key'])
            except Exception as e:
                print(f"Error updating Supabase Storage: {e}")
        
        return {"success": True, "message": "Quote updated successfully", "quote_id": quote_id}
    
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        import traceback
        print(f"Error updating quote: {e}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


# ========== ENDPOINTS DE PLANTILLAS ==========

@app.get("/plantillas")
async def get_plantillas(vendedor_id: str):
    """Obtiene todas las plantillas del vendedor"""
    conn = _get_connection()
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT id, nombre, descripcion, items_json, condiciones_ids,
                       plazo_dias, condicion_pago, veces_usada, created_at, updated_at
                FROM plantillas_cotizacion
                WHERE vendedor_id = %s AND activo = true
                ORDER BY veces_usada DESC, nombre ASC
            """, (vendedor_id,))
            plantillas = cur.fetchall()
            return [dict(p) for p in plantillas]
    except Exception as e:
        print(f"Error fetching plantillas: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@app.post("/plantillas")
async def create_plantilla(payload: dict):
    """Crea una nueva plantilla desde una cotización actual"""
    conn = _get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO plantillas_cotizacion 
                (nombre, descripcion, vendedor_id, items_json, condiciones_ids, 
                 plazo_dias, condicion_pago)
                VALUES (%s, %s, %s, %s, %s::uuid[], %s, %s)
                RETURNING id
            """, (
                payload.get('nombre'),
                payload.get('descripcion'),
                payload.get('vendedor_id'),
                json.dumps(payload.get('items'), ensure_ascii=False),
                payload.get('condiciones_ids', []),
                payload.get('plazo_dias'),
                payload.get('condicion_pago')
            ))
            plantilla_id = cur.fetchone()[0]
            conn.commit()
            return {"success": True, "plantilla_id": str(plantilla_id)}
    except Exception as e:
        conn.rollback()
        print(f"Error creating plantilla: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@app.get("/plantillas/{plantilla_id}")
async def get_plantilla(plantilla_id: str):
    """Obtiene una plantilla específica"""
    conn = _get_connection()
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT id, nombre, descripcion, items_json, condiciones_ids,
                       plazo_dias, condicion_pago, veces_usada
                FROM plantillas_cotizacion
                WHERE id = %s AND activo = true
            """, (plantilla_id,))
            plantilla = cur.fetchone()
            if not plantilla:
                raise HTTPException(status_code=404, detail="Plantilla no encontrada")
            
            # Incrementar contador de uso
            cur.execute("""
                UPDATE plantillas_cotizacion 
                SET veces_usada = veces_usada + 1 
                WHERE id = %s
            """, (plantilla_id,))
            conn.commit()
            
            return dict(plantilla)
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error fetching plantilla: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@app.put("/plantillas/{plantilla_id}")
async def update_plantilla(plantilla_id: str, payload: dict):
    """Actualiza una plantilla existente"""
    conn = _get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE plantillas_cotizacion
                SET nombre = %s, descripcion = %s, items_json = %s, 
                    condiciones_ids = %s::uuid[], plazo_dias = %s, condicion_pago = %s,
                    updated_at = NOW()
                WHERE id = %s
                RETURNING id
            """, (
                payload.get('nombre'),
                payload.get('descripcion'),
                json.dumps(payload.get('items'), ensure_ascii=False),
                payload.get('condiciones_ids', []),
                payload.get('plazo_dias'),
                payload.get('condicion_pago'),
                plantilla_id
            ))
            if not cur.fetchone():
                raise HTTPException(status_code=404, detail="Plantilla no encontrada")
            conn.commit()
            return {"success": True}
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        print(f"Error updating plantilla: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@app.delete("/plantillas/{plantilla_id}")
async def delete_plantilla(plantilla_id: str):
    """Elimina (desactiva) una plantilla"""
    conn = _get_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE plantillas_cotizacion
                SET activo = false, updated_at = NOW()
                WHERE id = %s
                RETURNING id
            """, (plantilla_id,))
            if not cur.fetchone():
                raise HTTPException(status_code=404, detail="Plantilla no encontrada")
            conn.commit()
            return {"success": True}
    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        print(f"Error deleting plantilla: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


# ========== OTROS ENDPOINTS ==========

@app.post("/quote/next-number")
async def get_next_quote_number():
    """Obtiene el siguiente número de cotización"""
    try:
        if _has_database_url():
            _ensure_sequence_table()
        
        year = date.today().year
        if _has_database_url():
            sequential = _next_quote_sequential(year)
            number = f"{sequential:03d}"
        else:
            number = "001"
        
        year_suffix = str(year)[-2:]
        token = f"{number}-{year_suffix}"
        return {"number": number, "year": year, "token": token}
    except Exception as e:
        import traceback
        print(f"Error in get_next_quote_number: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# ===================== USER PROFILE ENDPOINT =====================

@app.get("/user/me")
async def get_current_user(authorization: str = Header(None)):
    """Get current user profile from Directus token"""
    if not authorization:
        return {"data": None}
    
    try:
        # Forward request to Directus
        directus_url = os.getenv("DIRECTUS_URL", "http://directus:8055")
        resp = await asyncio.get_event_loop().run_in_executor(
            None,
            lambda: __import__('urllib.request', fromlist=['urlopen']).urlopen(
                __import__('urllib.request', fromlist=['Request']).Request(
                    f"{directus_url}/users/me",
                    headers={"Authorization": authorization}
                )
            )
        )
        import json as json_module
        data = json_module.loads(resp.read().decode())
        user = data.get('data', {})
        return {
            "data": {
                "id": user.get('id'),
                "first_name": user.get('first_name'),
                "last_name": user.get('last_name'),
                "email": user.get('email'),
                "phone": user.get('phone') or user.get('telefono'),
            }
        }
    except Exception as e:
        print(f"Error fetching user: {e}")
        return {"data": None}


# ===================== CLIENTS & PROJECTS ENDPOINTS =====================

@app.get("/clientes")
async def get_clientes(search: str = ""):
    """Get clients list with optional search - reads from CRM clientes table"""
    if not _has_database_url():
        return {"data": []}
    
    try:
        conn = _get_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            if search:
                cur.execute("""
                    SELECT id, nombre, email, telefono, empresa, estado, sector, ruc, direccion
                    FROM clientes 
                    WHERE (nombre ILIKE %s OR empresa ILIKE %s OR email ILIKE %s)
                    AND deleted_at IS NULL
                    ORDER BY nombre
                    LIMIT 20
                """, (f"%{search}%", f"%{search}%", f"%{search}%"))
            else:
                cur.execute("""
                    SELECT id, nombre, email, telefono, empresa, estado, sector, ruc, direccion
                    FROM clientes 
                    WHERE deleted_at IS NULL
                    ORDER BY nombre LIMIT 50
                """)
            results = cur.fetchall()
            # Map to cotizador expected format (B2B professional format)
            # empresa = company name (primary), nombre = contact person (secondary)
            mapped = [{
                'id': str(r['id']),
                'nombre': r.get('empresa') or r.get('nombre', ''),  # empresa as main client name
                'contacto': r.get('nombre', ''),  # nombre as contact person
                'email': r.get('email', ''),
                'telefono': r.get('telefono', ''),
                'ruc': r.get('ruc', ''),
                'direccion': r.get('direccion', ''),
            } for r in results]
            return {"data": mapped}
    except Exception as e:
        import traceback
        print(f"Error in get_clientes: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals() and conn:
            conn.close()


@app.post("/clientes")
async def create_cliente(data: dict):
    """Create a new client - uses same columns as CRM"""
    if not _has_database_url():
        raise HTTPException(status_code=400, detail="Database not configured")
    
    try:
        conn = _get_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                INSERT INTO clientes (nombre, email, telefono, empresa, ruc, estado, sector, direccion)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id, nombre, email, telefono, empresa, ruc, estado, sector, direccion
            """, (
                data.get('contacto', '') or data.get('nombre', ''),
                data.get('email', ''),
                data.get('telefono', ''),
                data.get('nombre', ''),
                data.get('ruc', ''),
                'prospecto',
                'General',
                data.get('direccion', '')
            ))
            result = cur.fetchone()
            conn.commit()
            # Map back to cotizador format
            mapped = {
                'id': str(result['id']),
                'nombre': result.get('empresa', ''),
                'contacto': result.get('nombre', ''),
                'email': result.get('email', ''),
                'telefono': result.get('telefono', ''),
                'ruc': result.get('ruc', ''),
                'direccion': result.get('direccion', ''),
            }
            return {"data": mapped}
    except Exception as e:
        import traceback
        print(f"Error in create_cliente: {e}")
        traceback.print_exc()
        if 'conn' in locals() and conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals() and conn:
            conn.close()


@app.get("/proyectos")
async def get_proyectos(cliente_id: str = None, search: str = ""):
    """Get projects list, optionally filtered by client - reads from CRM proyectos table"""
    if not _has_database_url():
        return {"data": []}
    
    try:
        conn = _get_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            query = """
                SELECT 
                    p.id, p.nombre, p.descripcion, p.cliente_id, p.created_at, p.direccion, p.ubicacion, 
                    c.empresa as cliente_nombre,
                    v.full_name as vendedor_nombre, v.phone as vendedor_telefono
                FROM proyectos p
                LEFT JOIN clientes c ON p.cliente_id = c.id
                LEFT JOIN vendedores v ON p.vendedor_id = v.id
                WHERE p.deleted_at IS NULL
            """
            params = []

            if cliente_id:
                query += " AND p.cliente_id = %s"
                params.append(cliente_id)
            
            if search:
                query += " AND p.nombre ILIKE %s"
                params.append(f"%{search}%")
            
            query += " ORDER BY p.nombre LIMIT 50"
            
            cur.execute(query, tuple(params))
            results = cur.fetchall()
            
            # Map results to ensure JSON serializability (handle datetime and UUID)
            mapped = []
            for r in results:
                mapped.append({
                    'id': str(r['id']),
                    'nombre': r['nombre'],
                    'direccion': r.get('direccion', ''),
                    'ubicacion': r.get('ubicacion', ''),
                    'descripcion': r.get('descripcion', ''),
                    'cliente_id': str(r['cliente_id']),
                    'cliente_nombre': r.get('cliente_nombre', ''),
                    'vendedor_nombre': r.get('vendedor_nombre', ''),
                    'vendedor_telefono': r.get('vendedor_telefono', ''),
                    'created_at': r['created_at'].isoformat() if r.get('created_at') else None
                })
            
            return {"data": mapped}
    except Exception as e:
        import traceback
        print(f"Error in get_proyectos: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals() and conn:
            conn.close()


@app.post("/proyectos")
async def create_proyecto(data: dict):
    """Create a new project (requires cliente_id)"""
    if not _has_database_url():
        raise HTTPException(status_code=400, detail="Database not configured")
    
    if not data.get('cliente_id'):
        raise HTTPException(status_code=400, detail="cliente_id is required")
    
    try:
        conn = _get_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            # Extract vendedor_id from data or fallback to None
            vendedor_id = data.get('vendedor_id') or data.get('user_id')
            
            cur.execute("""
                INSERT INTO proyectos (nombre, ubicacion, descripcion, cliente_id, vendedor_id)
                VALUES (%s, %s, %s, %s, %s)
                RETURNING id, nombre, ubicacion, descripcion, cliente_id, vendedor_id
            """, (
                data.get('nombre', ''),
                data.get('ubicacion', ''),
                data.get('descripcion', ''),
                data.get('cliente_id'),
                vendedor_id
            ))
            result = cur.fetchone()
            conn.commit()
            # Convert results to ensure JSON serializability
            mapped = {k: (str(v) if k in ('id', 'cliente_id', 'vendedor_id') else v) for k, v in result.items()}
            return {"data": mapped}
    except Exception as e:
        import traceback
        print(f"Error in create_proyecto: {e}")
        traceback.print_exc()
        if 'conn' in locals() and conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals() and conn:
            conn.close()


# ============================================================================
# CONDICIONES ESPECÍFICAS ENDPOINTS
# ============================================================================

@app.get("/condiciones")
async def get_condiciones(search: str = ""):
    """Get all active specific conditions, optionally filtered by search"""
    if not _has_database_url():
        return {"data": []}
    
    try:
        conn = _get_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            query = """
                SELECT id, texto, categoria, orden, created_by, created_at
                FROM condiciones_especificas
                WHERE activo = true
            """
            params = []
            
            if search:
                query += " AND texto ILIKE %s"
                params.append(f"%{search}%")
            
            query += " ORDER BY orden ASC, created_at ASC"
            
            cur.execute(query, params)
            results = cur.fetchall()
            # Ensure JSON serializable
            return {"data": [dict(r) for r in results]}
    except Exception as e:
        print(f"Error in get_condiciones: {e}")
        import traceback
        traceback.print_exc()
        return {"data": []}
    finally:
        if 'conn' in locals() and conn:
            conn.close()


@app.post("/condiciones")
async def create_condicion(data: dict):
    """Create a new specific condition"""
    if not _has_database_url():
        raise HTTPException(status_code=400, detail="Database not configured")
    
    try:
        conn = _get_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                INSERT INTO condiciones_especificas (texto, categoria, orden, created_by, activo)
                VALUES (%s, %s, %s, %s, %s)
                RETURNING id, texto, categoria, orden, created_by, created_at
            """, (
                data.get('texto', ''),
                data.get('categoria', ''),
                data.get('orden', 0),
                data.get('vendedor_id'),  # created_by = vendedor_id
                True
            ))
            result = cur.fetchone()
            conn.commit()
            return {"data": dict(result)}
    except Exception as e:
        import traceback
        print(f"Error in create_condicion: {e}")
        traceback.print_exc()
        if 'conn' in locals() and conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals() and conn:
            conn.close()


@app.put("/condiciones/{condicion_id}")
async def update_condicion(condicion_id: str, data: dict):
    """Update an existing condition"""
    if not _has_database_url():
        raise HTTPException(status_code=400, detail="Database not configured")
    
    try:
        conn = _get_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                UPDATE condiciones_especificas
                SET texto = %s, categoria = %s, orden = %s, updated_at = NOW()
                WHERE id = %s
                RETURNING id, texto, categoria, orden, created_at, updated_at
            """, (
                data.get('texto', ''),
                data.get('categoria', ''),
                data.get('orden', 0),
                condicion_id
            ))
            result = cur.fetchone()
            if not result:
                raise HTTPException(status_code=404, detail="Condición no encontrada")
            conn.commit()
            return {"data": dict(result)}
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print(f"Error in update_condicion: {e}")
        traceback.print_exc()
        if 'conn' in locals() and conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals() and conn:
            conn.close()


@app.delete("/condiciones/{condicion_id}")
async def delete_condicion(condicion_id: str):
    """Soft delete a condition (set activo = false)"""
    if not _has_database_url():
        raise HTTPException(status_code=400, detail="Database not configured")
    
    try:
        conn = _get_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                UPDATE condiciones_especificas
                SET activo = false, updated_at = NOW()
                WHERE id = %s
                RETURNING id
            """, (condicion_id,))
            result = cur.fetchone()
            if not result:
                raise HTTPException(status_code=404, detail="Condición no encontrada")
            conn.commit()
            return {"message": "Condición eliminada", "id": str(result['id'])}
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        print(f"Error in delete_condicion: {e}")
        traceback.print_exc()
        if 'conn' in locals() and conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals() and conn:
            conn.close()


# =====================================================
# ENDPOINTS: PROGRAMACIÓN DE SERVICIOS
# =====================================================

@app.get("/programacion")
async def get_programacion(
    seccion: str = None,  # 'laboratorio', 'comercial', 'administracion'
    estado: str = None,
    fecha_desde: str = None,
    fecha_hasta: str = None,
    search: str = None,
    limit: int = 100,
    offset: int = 0
):
    """Obtiene todos los registros de programación de servicios"""
    if not _has_database_url():
        return {"data": [], "total": 0}
    
    try:
        conn = _get_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            # Query base
            query = """
                SELECT * FROM programacion_servicios
                WHERE activo = true
            """
            params = []
            
            # Filtros opcionales
            if estado:
                query += " AND estado_trabajo = %s"
                params.append(estado)
            
            if fecha_desde:
                query += " AND fecha_recepcion >= %s"
                params.append(fecha_desde)
            
            if fecha_hasta:
                query += " AND fecha_recepcion <= %s"
                params.append(fecha_hasta)
            
            if search:
                query += """ AND (
                    cliente_nombre ILIKE %s OR
                    recep_numero ILIKE %s OR
                    ot ILIKE %s OR
                    cotizacion_lab ILIKE %s OR
                    proyecto ILIKE %s
                )"""
                search_param = f"%{search}%"
                params.extend([search_param] * 5)
            
            # Count total
            count_query = query.replace("SELECT *", "SELECT COUNT(*)")
            cur.execute(count_query, params)
            total = cur.fetchone()['count']
            
            # Order and pagination
            query += " ORDER BY item_numero DESC LIMIT %s OFFSET %s"
            params.extend([limit, offset])
            
            cur.execute(query, params)
            results = cur.fetchall()
            
            return {
                "data": [dict(r) for r in results],
                "total": total,
                "limit": limit,
                "offset": offset
            }
    except Exception as e:
        print(f"Error in get_programacion: {e}")
        import traceback
        traceback.print_exc()
        return {"data": [], "total": 0}
    finally:
        if 'conn' in locals() and conn:
            conn.close()


@app.get("/programacion/{prog_id}")
async def get_programacion_by_id(prog_id: str):
    """Obtiene un registro específico de programación"""
    if not _has_database_url():
        raise HTTPException(status_code=404, detail="Database not configured")
    
    try:
        conn = _get_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT * FROM programacion_servicios
                WHERE id = %s AND activo = true
            """, (prog_id,))
            result = cur.fetchone()
            if not result:
                raise HTTPException(status_code=404, detail="Registro no encontrado")
            return dict(result)
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in get_programacion_by_id: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals() and conn:
            conn.close()


@app.post("/programacion")
async def create_programacion(data: dict):
    """Crea un nuevo registro de programación (desde Laboratorio)"""
    if not _has_database_url():
        raise HTTPException(status_code=400, detail="Database not configured")
    
    try:
        conn = _get_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            # Generar recep_numero automático si no se proporciona
            recep_numero = data.get('recep_numero')
            if not recep_numero:
                cur.execute("SELECT generar_recep_numero() as numero")
                recep_numero = cur.fetchone()['numero']
            
            # Generar OT automático si no se proporciona
            ot = data.get('ot')
            if not ot:
                cur.execute("SELECT generar_ot_numero() as numero")
                ot = cur.fetchone()['numero']
            
            cur.execute("""
                INSERT INTO programacion_servicios (
                    recep_numero, ot, codigo_muestra, fecha_recepcion,
                    fecha_inicio, fecha_entrega_estimada, cliente_nombre,
                    descripcion_servicio, proyecto, estado_trabajo,
                    cotizacion_lab, created_by
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                )
                RETURNING *
            """, (
                recep_numero,
                ot,
                data.get('codigo_muestra'),
                data.get('fecha_recepcion'),
                data.get('fecha_inicio'),
                data.get('fecha_entrega_estimada'),
                data.get('cliente_nombre'),
                data.get('descripcion_servicio'),
                data.get('proyecto'),
                data.get('estado_trabajo', 'PROCESO'),
                data.get('cotizacion_lab'),
                data.get('user_id')
            ))
            result = cur.fetchone()
            conn.commit()
            return {"data": dict(result), "message": "Registro creado exitosamente"}
    except Exception as e:
        print(f"Error in create_programacion: {e}")
        import traceback
        traceback.print_exc()
        if 'conn' in locals() and conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals() and conn:
            conn.close()


@app.put("/programacion/{prog_id}/laboratorio")
async def update_programacion_laboratorio(prog_id: str, data: dict):
    """Actualiza campos de la sección LABORATORIO"""
    if not _has_database_url():
        raise HTTPException(status_code=400, detail="Database not configured")
    
    allowed_fields = [
        'ot', 'codigo_muestra', 'fecha_recepcion', 'fecha_inicio',
        'fecha_entrega_estimada', 'cliente_nombre', 'descripcion_servicio',
        'proyecto', 'entrega_real', 'estado_trabajo', 'cotizacion_lab',
        'autorizacion_lab', 'nota_lab', 'dias_atraso_lab',
        'motivo_dias_atraso_lab', 'evidencia_envio_recepcion', 'envio_informes'
    ]
    
    try:
        conn = _get_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            # Construir query dinámico solo con campos permitidos
            updates = []
            params = []
            for field in allowed_fields:
                if field in data:
                    updates.append(f"{field} = %s")
                    params.append(data[field])
            
            if not updates:
                raise HTTPException(status_code=400, detail="No hay campos para actualizar")
            
            # Agregar updated_by
            updates.append("updated_by = %s")
            params.append(data.get('user_id'))
            
            params.append(prog_id)
            
            query = f"""
                UPDATE programacion_servicios
                SET {', '.join(updates)}
                WHERE id = %s AND activo = true
                RETURNING *
            """
            
            cur.execute(query, params)
            result = cur.fetchone()
            
            if not result:
                raise HTTPException(status_code=404, detail="Registro no encontrado")
            
            conn.commit()
            return {"data": dict(result), "message": "Sección Laboratorio actualizada"}
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in update_programacion_laboratorio: {e}")
        import traceback
        traceback.print_exc()
        if 'conn' in locals() and conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals() and conn:
            conn.close()


@app.put("/programacion/{prog_id}/comercial")
async def update_programacion_comercial(prog_id: str, data: dict):
    """Actualiza campos de la sección COMERCIAL"""
    if not _has_database_url():
        raise HTTPException(status_code=400, detail="Database not configured")
    
    allowed_fields = [
        'fecha_solicitud_com', 'fecha_entrega_com',
        'evidencia_solicitud_envio', 'dias_atraso_envio_coti',
        'motivo_dias_atraso_com'
    ]
    
    try:
        conn = _get_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            updates = []
            params = []
            for field in allowed_fields:
                if field in data:
                    updates.append(f"{field} = %s")
                    params.append(data[field])
            
            if not updates:
                raise HTTPException(status_code=400, detail="No hay campos para actualizar")
            
            updates.append("updated_by = %s")
            params.append(data.get('user_id'))
            params.append(prog_id)
            
            query = f"""
                UPDATE programacion_servicios
                SET {', '.join(updates)}
                WHERE id = %s AND activo = true
                RETURNING *
            """
            
            cur.execute(query, params)
            result = cur.fetchone()
            
            if not result:
                raise HTTPException(status_code=404, detail="Registro no encontrado")
            
            conn.commit()
            return {"data": dict(result), "message": "Sección Comercial actualizada"}
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in update_programacion_comercial: {e}")
        import traceback
        traceback.print_exc()
        if 'conn' in locals() and conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals() and conn:
            conn.close()


@app.put("/programacion/{prog_id}/administracion")
async def update_programacion_administracion(prog_id: str, data: dict):
    """Actualiza campos de la sección ADMINISTRACIÓN"""
    if not _has_database_url():
        raise HTTPException(status_code=400, detail="Database not configured")
    
    allowed_fields = [
        'numero_factura', 'estado_pago', 'estado_autorizar', 'nota_admin'
    ]
    
    try:
        conn = _get_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            updates = []
            params = []
            for field in allowed_fields:
                if field in data:
                    updates.append(f"{field} = %s")
                    params.append(data[field])
            
            if not updates:
                raise HTTPException(status_code=400, detail="No hay campos para actualizar")
            
            updates.append("updated_by = %s")
            params.append(data.get('user_id'))
            params.append(prog_id)
            
            query = f"""
                UPDATE programacion_servicios
                SET {', '.join(updates)}
                WHERE id = %s AND activo = true
                RETURNING *
            """
            
            cur.execute(query, params)
            result = cur.fetchone()
            
            if not result:
                raise HTTPException(status_code=404, detail="Registro no encontrado")
            
            conn.commit()
            return {"data": dict(result), "message": "Sección Administración actualizada"}
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in update_programacion_administracion: {e}")
        import traceback
        traceback.print_exc()
        if 'conn' in locals() and conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals() and conn:
            conn.close()


@app.delete("/programacion/{prog_id}")
async def delete_programacion(prog_id: str):
    """Elimina (soft delete) un registro de programación"""
    if not _has_database_url():
        raise HTTPException(status_code=400, detail="Database not configured")
    
    try:
        conn = _get_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                UPDATE programacion_servicios
                SET activo = false, updated_at = NOW()
                WHERE id = %s
                RETURNING id
            """, (prog_id,))
            result = cur.fetchone()
            
            if not result:
                raise HTTPException(status_code=404, detail="Registro no encontrado")
            
            conn.commit()
            return {"message": "Registro eliminado", "id": str(result['id'])}
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in delete_programacion: {e}")
        if 'conn' in locals() and conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if 'conn' in locals() and conn:
            conn.close()


@app.get("/programacion/next-numbers")
async def get_next_numbers():
    """Obtiene los próximos números de RECEP y OT"""
    if not _has_database_url():
        return {"recep_numero": "1-26", "ot": "1-26 LEM"}
    
    try:
        conn = _get_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("SELECT generar_recep_numero() as recep, generar_ot_numero() as ot")
            result = cur.fetchone()
            return {
                "recep_numero": result['recep'],
                "ot": result['ot']
            }
    except Exception as e:
        print(f"Error getting next numbers: {e}")
        return {"recep_numero": "1-26", "ot": "1-26 LEM"}
    finally:
        if 'conn' in locals() and conn:
            conn.close()


@app.get("/programacion/{prog_id}/historial")
async def get_programacion_historial(prog_id: str):
    """Obtiene el historial de cambios de un registro"""
    if not _has_database_url():
        return {"data": []}
    
    try:
        conn = _get_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT h.*, v.full_name as modificado_por_nombre
                FROM programacion_servicios_historial h
                LEFT JOIN vendedores v ON h.modificado_por = v.id
                WHERE h.programacion_id = %s
                ORDER BY h.modificado_at DESC
                LIMIT 50
            """, (prog_id,))
            results = cur.fetchall()
            return {"data": [dict(r) for r in results]}
    except Exception as e:
        print(f"Error in get_programacion_historial: {e}")
        return {"data": []}
    finally:
        if 'conn' in locals() and conn:
            conn.close()


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="0.0.0.0", port=8000)
